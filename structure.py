import os
from openpyxl import Workbook
from openpyxl.styles import Font

ROOT_DIR = "."
OUTPUT_FILE = "project_structure.xlsx"

EXCLUDE_DIRS = {
    ".venv", "venv", "__pycache__", ".git",
    ".idea", ".vscode", "node_modules",
    "dist", "build"
}

EXCLUDE_EXTENSIONS = {
    ".pyc", ".log", ".tmp",
    ".jpg", ".jpeg", ".png"
}

wb = Workbook()
ws = wb.active
ws.title = "Project Structure"

row = 1
FONT = Font(name="Consolas")

def write_tree(path, prefix="", is_last=True):
    global row

    name = os.path.basename(path)
    connector = "└─ " if is_last else "├─ "
    line = f"{prefix}{connector}"

    if os.path.isdir(path):
        line += f"📁 {name}"
    else:
        line += f"📄 {name}"

    ws.cell(row=row, column=1, value=line)
    ws.cell(row=row, column=1).font = FONT
    row += 1

    if not os.path.isdir(path):
        return

    try:
        items = sorted(os.listdir(path))
    except PermissionError:
        return

    # lọc item
    filtered = []
    for item in items:
        if item in EXCLUDE_DIRS:
            continue
        full = os.path.join(path, item)
        if os.path.isfile(full) and any(item.lower().endswith(ext) for ext in EXCLUDE_EXTENSIONS):
            continue
        filtered.append(item)

    for i, item in enumerate(filtered):
        full_path = os.path.join(path, item)
        last = i == len(filtered) - 1
        new_prefix = prefix + ("   " if is_last else "│  ")
        write_tree(full_path, new_prefix, last)

# ROOT
root_name = os.path.basename(os.path.abspath(ROOT_DIR))
ws.cell(row=row, column=1, value=f"📁 {root_name}")
ws.cell(row=row, column=1).font = Font(name="Consolas", bold=True)
row += 1

items = sorted(os.listdir(ROOT_DIR))
items = [
    i for i in items
    if i not in EXCLUDE_DIRS
    and not (os.path.isfile(i) and any(i.lower().endswith(ext) for ext in EXCLUDE_EXTENSIONS))
]

for i, item in enumerate(items):
    write_tree(
        os.path.join(ROOT_DIR, item),
        prefix="",
        is_last=(i == len(items) - 1)
    )

ws.column_dimensions["A"].width = 140
wb.save(OUTPUT_FILE)

print(f"✅ Đã xuất cấu trúc thư mục dạng tree ra Excel: {OUTPUT_FILE}")
